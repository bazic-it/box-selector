import csv
import openpyxl
from utils import *
from config import *
from classes import *
from packer_main import Packer, Bin, Item

def registerBoxesToPacker(packer, boxes):
    for box in boxes:
        bin = Bin(box.name, box.length, box.width, box.height, MAX_WEIGHT_PER_BOX)
        packer.add_bin(bin)

def registerItemsToPacker(packer, items):
    for _item in items:
        item = Item(_item.sku, _item.uomCode, _item.length, _item.width, _item.height, _item.weight)
        packer.add_item(item)

def sortOrders(a, b):
    if a[2] == 'CASE' and (b[2] == 'BOX' or b[2] == 'EA'):
        return -1
    elif a[2] == 'BOX' and b[2] == 'EA':
        return -1
    else:
        return 1

def sortBoxes(boxes):
    return sorted(boxes, key=lambda b:b['volume'])

def getInventoryMasterData(inputFilepath):
    message = None
    targetColumns = ['Item No.', 'Available Qty', 'Case Length', 'Case Width', 'Case Height', 'Case Volume', 'Case Weight', 'Box Length', 'Box Width', 'Box Height', 'Box Volume', 'Box Weight', 'EA Length', 'EA Width', 'EA Height', 'EA Volume', 'EA Weight']
    keyColumn = 'Item No.'
    headerMap = {}
    mapped = {}

    try:
        age = getDaysDifferent(getCurrentime(), getFileModifiedDate(inputFilepath))
        message = 'Inventory master file was updated {} days ago.'.format(age)

        workbook = openpyxl.load_workbook(inputFilepath)
        sheet = workbook.active
        for r in range(1, sheet.max_row+1):
            itemNumber = None
            for c in range(1, sheet.max_column+1):
                data = sheet.cell(row=r, column=c).value
                if r == 1:
                    for colName in targetColumns:
                        if data == colName:
                            headerMap[c] = colName
                else:
                    if c in headerMap:
                        if headerMap[c] == keyColumn:
                            itemNumber = str(data)
                            mapped[itemNumber] = {}
                        else:
                            mapped[itemNumber][headerMap[c]] = data
    except Exception as e:
        print('*** Error: Failed to read input file for Inventory Master: {} ***'.format(e))
        return {}, message
    
    return mapped, message

def getSalesQuotationItemsFromInputfile(filepath):
    message = None
    targetColumns = ['Item No.', 'Item Description', 'UoM Code', 'Quantity', 'Price Per Piece', 'Total (LC)', 'Unit Price', 'Available Qty']
    headerMap = {}
    items = []
    
    try:
        age = getDaysDifferent(getCurrentime(), getFileModifiedDate(filepath))
        message = 'Inventory master file was updated {} days ago.'.format(age)

        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        for r in range(1, sheet.max_row+1):
            item = {}
            for c in range(1, sheet.max_column+1):
                data = sheet.cell(row=r, column=c).value
                if r == 1:
                    for colName in targetColumns:
                        if data == colName:
                            headerMap[c] = colName
                else:
                    if c in headerMap:
                        item[headerMap[c]] = data
            if item:
                items.append(item)

    except Exception as e:
        print('*** Error: Failed to read input file for Sales Quotation: {} ***'.format(e))
        return {}, message
    
    return items, message

def getBoxesMasterData(inputFilepath):
    boxes = []
    message = None

    try:
        age = getDaysDifferent(getCurrentime(), getFileModifiedDate(inputFilepath))
        message = 'Boxes master file was updated {} days ago.'.format(age)
        row = 0
        with open (inputFilepath, mode='r') as file:
            content = csv.reader(file)
            for line in content:
                if row == 0:
                    row += 1
                    continue
                length = float(line[1]) - BOX_DIMENSION_PADDING
                width = float(line[2]) - BOX_DIMENSION_PADDING
                height = float(line[3]) - BOX_DIMENSION_PADDING
                weight = float(line[4])
                box = Box(line[0], length, width, height, weight)
                boxes.append(box)
                row += 1
    except Exception as e:
        print(f'*** Error: Failed to read input file for Boxes Master Data. Please make sure filename is valid. {e} ***')
        return {}, message

    return boxes, message

def convertStringToFloat(value):
    if type(value) == str:
        value = value.replace('$', '')
        return float(value)
    return float(value)

def combineDetailsForEachItem(inventoryMaster, items):
    itemLines = []
    itemsWithNoInfo = []

    for i in items:
        if not i['Item No.']:
            continue
        sku = i['Item No.']
        if sku in inventoryMaster:
            item = ItemLine()
            item.sku = sku
            item.itemDescription = i['Item Description']
            item.qty = int(i['Quantity'])
            item.pricePerPiece = convertStringToFloat(i['Price Per Piece'])
            item.totalLC = convertStringToFloat(i['Total (LC)'])
            item.unitPrice = convertStringToFloat(i['Unit Price'])
            item.available = int(i['Available Qty'])

            uomCode = i['UoM Code']
            if uomCode:
                item.uomCode = uomCode
                if uomCode == 'EA':
                    item.length = float(inventoryMaster[sku]['EA Length']) if 'EA Length' in inventoryMaster[sku] else None
                    item.width = float(inventoryMaster[sku]['EA Width']) if 'EA Width' in inventoryMaster[sku] else None
                    item.height = float(inventoryMaster[sku]['EA Height']) if 'EA Height' in inventoryMaster[sku] else None
                    item.volume = float(inventoryMaster[sku]['EA Volume']) if 'EA Volume' in inventoryMaster[sku] else None
                    item.weight = float(inventoryMaster[sku]['EA Weight']) if 'EA Weight' in inventoryMaster[sku] else None
                elif uomCode == 'BOX':
                    item.length = float(inventoryMaster[sku]['Box Length']) if 'Box Length' in inventoryMaster[sku] else None
                    item.width = float(inventoryMaster[sku]['Box Width']) if 'Box Width' in inventoryMaster[sku] else None
                    item.height = float(inventoryMaster[sku]['Box Height']) if 'Box Height' in inventoryMaster[sku] else None
                    item.volume = float(inventoryMaster[sku]['Box Volume']) if 'Box Volume' in inventoryMaster[sku] else None
                    item.weight = float(inventoryMaster[sku]['Box Weight']) if 'Box Weight' in inventoryMaster[sku] else None
                elif uomCode == 'CASE':
                    item.length = float(inventoryMaster[sku]['Case Length']) if 'Case Length' in inventoryMaster[sku] else None
                    item.width = float(inventoryMaster[sku]['Case Width']) if 'Case Width' in inventoryMaster[sku] else None
                    item.height = float(inventoryMaster[sku]['Case Height']) if 'Case Height' in inventoryMaster[sku] else None
                    item.volume = float(inventoryMaster[sku]['Case Volume']) if 'Case Volume' in inventoryMaster[sku] else None
                    item.weight = float(inventoryMaster[sku]['Case Weight']) if 'Case Weight' in inventoryMaster[sku] else None
                else:
                    item.length = None
                    item.width = None
                    item.height = None
                    item.volume = None
                    item.weight = None
            itemLines.append(item)
        else:
            itemsWithNoInfo.append(sku)
    
    return itemLines, itemsWithNoInfo

def splitItem(itemLines):
    newItemLines = []

    for item in itemLines:
        # if item.uomCode == 'CASE' and item.qty > 1:
        if item.qty > 1:
            itemQty = item.qty
            item.qty = 1
            newItemLines.append(item)
            for _ in range(itemQty - 1):
                newItem = ItemLine()
                newItem.sku = item.sku
                newItem.itemDescription = item.itemDescription
                newItem.uomCode = item.uomCode
                newItem.qty = 1
                newItem.pricePerPiece = item.pricePerPiece
                newItem.totalLC = item.totalLC
                newItem.unitPrice = item.unitPrice
                newItem.available = item.available
                newItem.length = item.length
                newItem.width = item.width
                newItem.height = item.height
                newItem.volume = item.volume
                newItem.weight = item.weight
                newItemLines.append(item)
        else:
            newItemLines.append(item)

    return newItemLines

def itemFitByDimension(boxLength, boxWidth, boxHeight, itemLength, itemWidth, itemHeight):
    failCondition1 = itemLength > boxLength or itemWidth > boxWidth or itemHeight > boxHeight
    failCondition2 = itemWidth > boxLength or itemHeight > boxWidth or itemLength > boxHeight
    failCondition3 = itemHeight > boxLength or itemLength > boxWidth or itemWidth > boxHeight

    if failCondition1 and failCondition2 and failCondition3:
        return False
    return True

def compileItemsInBox(items):
    mapped = {}

    for item in items:
        key = item.name + "-" + item.uom
        if key in mapped:
            mapped[key]["qty"] += 1
        else:
            mapped[key] = {
                "name": item.name,
                "uom": item.uom,
                "qty": 1,
                "width": item.width,
                "height": item.height,
                "depth": item.depth
            }

    return mapped

def displayResultsAsString(boxes, itemsWithoutOuterBox):
    texts = []
    count = 1

    for box in boxes:
        if box.items:
            compiledItems = compileItemsInBox(box.items)
            texts.append('{}. {}    ({}" x {}" x {}")'.format(count, box.name, float(box.width) + BOX_DIMENSION_PADDING, float(box.height) + BOX_DIMENSION_PADDING, float(box.depth) + BOX_DIMENSION_PADDING))
            texts.append('    Volume: {:.2f} / {:.2f} in³  Weight: {} Lbs'.format(box.get_filled_volume(), box.get_volume(), box.current_weight if box.current_weight > 0 else 1))
            texts.append('    Content:')
            for _, item in compiledItems.items():
                texts.append('      {}x   {} - {} ({}" x {}" x {}")'.format(item["qty"], item["name"], item["uom"], item["width"], item["height"], item["depth"]))
            texts.append('')
            count += 1

    if itemsWithoutOuterBox:
        texts.append('Ship As Is:')
        for item in itemsWithoutOuterBox:
            texts.append('      {} - {} ({}" x {}" x {}")'.format(item.name, item.uom, item.width, item.height, item.depth))

    return texts

def distribute(filepath):
    packer = Packer()
    success = True

    salesQuotationFilepath = validateInputFilename(filepath)

    inventoryMaster, invMsg = getInventoryMasterData(getInventoryMasterFilepath())
    boxesMaster, boxMsg = getBoxesMasterData(getBoxMasterFilepath())
    items, itemsMsg = getSalesQuotationItemsFromInputfile(salesQuotationFilepath)
    itemLines, itemsWithNoInfo = combineDetailsForEachItem(inventoryMaster, items)
    splittedItemLines = splitItem(itemLines)

    registerItemsToPacker(packer, splittedItemLines)
    registerBoxesToPacker(packer, boxesMaster)

    leftoverItems = packer.pack(bins_bigger_first=False, items_bigger_first=True, distribute_items=True, number_of_decimals=2)

    results = displayResultsAsString(packer.filled_bins, leftoverItems)

    return {
        'success': success,
        'results': results
    }

def validateInputFilename(filename):
    cleaned = filename
    if '/' in filename:
        cleaned = filename.split('/')[-1]

    if '.xlsx' not in cleaned:
        cleaned = cleaned + '.xlsx'

    return USER_DOWNLOADS + cleaned

# def writeLog(timestamp, status):
#     path = os.path.join(ASSETS_BASE_DIR, LOGS_FILENAME)
#     user = os.getenv('COMPUTERNAME')
#     try:
#         with open(path, 'a') as file:
#             file.write('USR;{} | IN;{} | SUCCESS;{} | ERR;{} | WARNING;{} | WARN;{} | OOS;{} | OUT;{} | VER;{} | TS;{}\n'.format(user, status["inputFilename"], status["success"], status["errorMessage"], status["warning"], status["warningMessage"], status["outOfStockSKUs"], status["outputFilename"], APP_VERSION, timestamp))
#     except:
#         print('*** Error: Failed to write to logs. ***')